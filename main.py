import html
import os
import traceback
from datetime import date, timedelta

import requests
import yfinance as yf
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

load_dotenv()

VERDE = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
NARANJA = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

TICKERS = [
    'GGAL.BA', 'YPFD.BA', 'PAMP.BA', 'BMA.BA', 'TXAR.BA',
    'ALUA.BA', 'CEPU.BA', 'EDN.BA', 'LOMA.BA', 'TGSU2.BA',
]


def fetch_weekly_data(tickers):
    """Baja velas semanales y devuelve la última vela cerrada por ticker.

    El bot corre los viernes después del cierre del mercado. A esa altura la
    vela de la semana en curso ya es definitiva (no quedan ruedas hasta el
    lunes), así que entra en el reporte. De lunes a jueves se la descarta, por
    si alguien dispara una corrida manual a mitad de semana con datos a medias.
    """
    today = date.today()
    current_monday = today - timedelta(days=today.weekday())

    # yfinance indexa cada vela semanal por el lunes de su semana. El límite
    # define hasta qué vela se considera cerrada: los viernes/fin de semana
    # incluye la semana en curso; de lunes a jueves solo las semanas pasadas.
    if today.weekday() >= 4:  # viernes, sábado o domingo
        limite = current_monday + timedelta(days=1)
    else:
        limite = current_monday

    data = {}
    for ticker in tickers:
        # Cada ticker se aísla: si yfinance falla en uno (red, rate limit,
        # datos raros) el resto del reporte igual sale.
        try:
            df = yf.Ticker(ticker).history(period='3mo', interval='1wk')

            if df.empty:
                data[ticker] = None
                continue

            closed = df[[d.date() < limite for d in df.index]]
            if closed.empty:
                data[ticker] = None
                continue

            last = closed.iloc[-1]
            data[ticker] = {
                'fecha': closed.index[-1].date().isoformat(),
                'open': round(float(last['Open']), 2),
                'high': round(float(last['High']), 2),
                'low': round(float(last['Low']), 2),
                'close': round(float(last['Close']), 2),
            }
        except Exception:
            data[ticker] = None

    # Si NINGÚN ticker trajo datos, algo serio falló: fallar fuerte para que
    # el admin se entere, en vez de mandar un reporte vacío como si nada.
    if not any(data.values()):
        raise RuntimeError('yfinance no devolvió datos para ningún ticker')

    return data


def _calcular_alcistas(ws, n):
    """Lista los tickers cuyo cierre de la semana n superó al de la n-1.

    Cada alcista trae un flag 'barrida' (toma de liquidez): es True cuando,
    además de cerrar arriba, el mínimo de la semana n perforó al de la n-1.
    """
    if n < 2:
        return []

    col_close = 1 + 5 * n
    col_close_prev = 1 + 5 * (n - 1)
    col_min = col_close - 1
    col_min_prev = col_close_prev - 1

    alcistas = []
    for fila in range(2, ws.max_row + 1):
        ticker = ws.cell(row=fila, column=1).value
        cierre = ws.cell(row=fila, column=col_close).value
        cierre_prev = ws.cell(row=fila, column=col_close_prev).value
        if not ticker:
            continue
        if isinstance(cierre, (int, float)) and isinstance(cierre_prev, (int, float)):
            if cierre > cierre_prev:
                variacion = (cierre - cierre_prev) / cierre_prev * 100
                minimo = ws.cell(row=fila, column=col_min).value
                minimo_prev = ws.cell(row=fila, column=col_min_prev).value
                barrida = (
                    isinstance(minimo, (int, float))
                    and isinstance(minimo_prev, (int, float))
                    and minimo < minimo_prev
                )
                alcistas.append({
                    'ticker': ticker,
                    'variacion': round(variacion, 2),
                    'barrida': barrida,
                })
    return alcistas


def update_excel(data, filename='registro_merval.xlsx'):
    """Registra la semana en el Excel y devuelve los alcistas.

    Cada corrida suma un bloque de 5 columnas (fecha + OHLC). Es idempotente:
    si la última semana registrada ya es la de estos datos, no agrega nada,
    así un re-run (manual, reintento de Actions) no duplica columnas.
    """
    fecha_actual = next((v['fecha'] for v in data.values() if v), None)

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Ticker'
        for i, ticker in enumerate(data, start=2):
            ws.cell(row=i, column=1, value=ticker)

    semanas_previas = (ws.max_column - 1) // 5

    # Idempotencia: si la última semana ya tiene esta fecha, no duplicar el
    # bloque. Se recalculan los alcistas desde lo que ya está escrito.
    if semanas_previas >= 1 and fecha_actual:
        col_fecha_ultima = 2 + 5 * (semanas_previas - 1)
        fechas_ultima = {
            ws.cell(row=f, column=col_fecha_ultima).value
            for f in range(2, ws.max_row + 1)
        }
        if fecha_actual in fechas_ultima:
            return _calcular_alcistas(ws, semanas_previas)

    n = semanas_previas + 1
    base = 1 + 5 * semanas_previas  # última columna ocupada

    encabezados = [
        (base + 1, f'Semana {n} - Fecha'),
        (base + 2, f'Semana {n} - Apertura'),
        (base + 3, f'Semana {n} - Máximo'),
        (base + 4, f'Semana {n} - Mínimo'),
        (base + 5, f'Semana {n} - Cierre'),
    ]
    for col, texto in encabezados:
        ws.cell(row=1, column=col, value=texto)

    # Mapa ticker -> fila a partir de la columna A
    filas = {}
    for fila in range(2, ws.max_row + 1):
        nombre = ws.cell(row=fila, column=1).value
        if nombre:
            filas[nombre] = fila

    for ticker, vela in data.items():
        if vela is None or ticker not in filas:
            continue
        fila = filas[ticker]
        ws.cell(row=fila, column=base + 1, value=vela['fecha'])
        ws.cell(row=fila, column=base + 2, value=vela['open'])
        ws.cell(row=fila, column=base + 3, value=vela['high'])
        ws.cell(row=fila, column=base + 4, value=vela['low'])
        ws.cell(row=fila, column=base + 5, value=vela['close'])

    alcistas = _calcular_alcistas(ws, n)
    for item in alcistas:
        fila = filas[item['ticker']]
        ws.cell(row=fila, column=base + 5).fill = VERDE
        if item['barrida']:
            ws.cell(row=fila, column=base + 4).fill = NARANJA

    wb.save(filename)
    return alcistas


def build_message(data, alcistas):
    """Arma el mensaje en HTML para Telegram con el cierre semanal."""
    fecha = next((v['fecha'] for v in data.values() if v), 'sin datos')

    lineas = [
        '📊 <b>Cierre semanal Merval</b>',
        f'Semana del {fecha}',
        '',
    ]

    if alcistas:
        lineas.append(
            '🟢 <b>ATENCIÓN</b> — estas acciones cerraron por encima de la semana pasada:'
        )
        for item in alcistas:
            vela = data.get(item['ticker'])
            cierre = vela['close'] if vela else '-'
            fuego = ' 🔥' if item['barrida'] else ''
            lineas.append(
                f"• <b>{item['ticker']}</b> — ${cierre} (+{item['variacion']}%){fuego}"
            )
        if any(item['barrida'] for item in alcistas):
            lineas.append('')
            lineas.append(
                '🔥 toma de liquidez: la acción perforó el mínimo de la semana '
                'anterior y aun así cerró arriba. El mercado podría estar girando '
                'al alza.'
            )
    else:
        lineas.append('Ninguna acción cerró alcista esta semana')

    lineas.append('')
    lineas.append('<b>Resumen de cierres</b>')
    for ticker, vela in data.items():
        cierre = vela['close'] if vela else 'sin datos'
        lineas.append(f'• {ticker} — ${cierre}')

    return '\n'.join(lineas)


def send_telegram(message, chat_id):
    """Envía un mensaje HTML al chat de Telegram indicado."""
    token = os.getenv('TELEGRAM_TOKEN')

    if not token or not chat_id:
        raise RuntimeError('Faltan TELEGRAM_TOKEN o el chat_id de destino')

    url = f'https://api.telegram.org/bot{token}/sendMessage'
    resp = requests.post(
        url,
        data={'chat_id': chat_id, 'text': message, 'parse_mode': 'HTML'},
        timeout=15,
    )
    # No usar raise_for_status(): su mensaje incluye la URL completa, y la URL
    # lleva el token del bot. Eso terminaría en los logs de Actions y en el
    # aviso de error a Telegram. Se arma un error propio sin el token.
    if not resp.ok:
        raise RuntimeError(
            f'Telegram respondió {resp.status_code} al enviar a {chat_id}: '
            f'{resp.text[:300]}'
        )
    return resp.json()


def notify_all(message):
    """Manda el reporte al destinatario principal y al admin si está configurado."""
    chat_id = os.getenv('CHAT_ID')
    chat_id_admin = os.getenv('CHAT_ID_ADMIN')

    if not chat_id:
        raise RuntimeError('Falta CHAT_ID en el entorno')

    send_telegram(message, chat_id)
    if chat_id_admin and chat_id_admin != chat_id:
        send_telegram(message, chat_id_admin)


def main():
    try:
        data = fetch_weekly_data(TICKERS)
        alcistas = update_excel(data)
        message = build_message(data, alcistas)

        print(message)
        notify_all(message)
    except Exception:
        chat_id_admin = os.getenv('CHAT_ID_ADMIN')
        if chat_id_admin:
            tb = traceback.format_exc()

            # Si el token aparece en el traceback, borrarlo antes de mandarlo.
            token = os.getenv('TELEGRAM_TOKEN')
            if token:
                tb = tb.replace(token, '***')

            # Telegram corta a 4096 caracteres y el parser HTML rechaza '<' y
            # '&' sueltos. Sin escapar ni recortar, el aviso de error fallaría
            # justo cuando más se lo necesita. Se toma el final del traceback,
            # que es donde está el error real.
            tb = html.escape(tb[-3000:])
            aviso = f'⚠️ <b>Error en merval-bot</b>\n<pre>{tb}</pre>'

            try:
                send_telegram(aviso, chat_id_admin)
            except Exception:
                pass  # si falla el aviso, no tapar el error original
        raise


if __name__ == '__main__':
    main()
