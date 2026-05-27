"""Tests sobre update_excel y _calcular_alcistas.

Cubren el contrato crítico del Excel: idempotencia, detección de alcistas y
del patrón "toma de liquidez" (barrida del mínimo previo con cierre arriba).
"""

import os

import pytest
from openpyxl import load_workbook

import main


@pytest.fixture
def excel_path(tmp_path):
    return str(tmp_path / "registro.xlsx")


def _vela(fecha, open_, high, low, close):
    return {"fecha": fecha, "open": open_, "high": high, "low": low, "close": close}


def test_update_excel_crea_archivo_nuevo_con_encabezados(excel_path):
    data = {"GGAL.BA": _vela("2026-05-18", 100.0, 110.0, 99.0, 108.0)}

    main.update_excel(data, filename=excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws["A1"].value == "Ticker"
    assert ws["A2"].value == "GGAL.BA"
    assert ws.cell(row=1, column=2).value == "Semana 1 - Fecha"
    assert ws.cell(row=2, column=6).value == 108.0


def test_update_excel_es_idempotente(excel_path):
    data = {"GGAL.BA": _vela("2026-05-18", 100.0, 110.0, 99.0, 108.0)}

    main.update_excel(data, filename=excel_path)
    main.update_excel(data, filename=excel_path)  # mismo run repetido
    main.update_excel(data, filename=excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    # 1 columna ticker + 5 columnas por semana = 6 columnas, NO 16
    assert ws.max_column == 6


def test_update_excel_agrega_bloque_nuevo_en_segunda_semana(excel_path):
    main.update_excel(
        {"GGAL.BA": _vela("2026-05-18", 100.0, 110.0, 99.0, 108.0)},
        filename=excel_path,
    )
    main.update_excel(
        {"GGAL.BA": _vela("2026-05-25", 108.0, 115.0, 105.0, 112.0)},
        filename=excel_path,
    )

    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws.max_column == 11  # 1 + 5*2
    assert ws.cell(row=1, column=7).value == "Semana 2 - Fecha"
    assert ws.cell(row=2, column=11).value == 112.0


def test_update_excel_detecta_alcista_y_devuelve_variacion(excel_path):
    main.update_excel(
        {"GGAL.BA": _vela("2026-05-18", 100.0, 110.0, 99.0, 100.0)},
        filename=excel_path,
    )
    alcistas = main.update_excel(
        {"GGAL.BA": _vela("2026-05-25", 100.0, 115.0, 95.0, 110.0)},
        filename=excel_path,
    )

    assert len(alcistas) == 1
    assert alcistas[0]["ticker"] == "GGAL.BA"
    assert alcistas[0]["variacion"] == 10.0


def test_update_excel_detecta_toma_de_liquidez(excel_path):
    """Cierre arriba + mínimo perforando el de la semana previa = barrida."""
    main.update_excel(
        {"GGAL.BA": _vela("2026-05-18", 100.0, 110.0, 99.0, 100.0)},
        filename=excel_path,
    )
    alcistas = main.update_excel(
        # mínimo 95 < mínimo previo 99, cierre 110 > 100 → barrida True
        {"GGAL.BA": _vela("2026-05-25", 100.0, 115.0, 95.0, 110.0)},
        filename=excel_path,
    )

    assert alcistas[0]["barrida"] is True


def test_update_excel_no_marca_barrida_si_minimo_no_perfora(excel_path):
    main.update_excel(
        {"GGAL.BA": _vela("2026-05-18", 100.0, 110.0, 99.0, 100.0)},
        filename=excel_path,
    )
    alcistas = main.update_excel(
        # mínimo 100 > mínimo previo 99 → NO hay barrida pese a cerrar arriba
        {"GGAL.BA": _vela("2026-05-25", 100.0, 115.0, 100.0, 110.0)},
        filename=excel_path,
    )

    assert alcistas[0]["barrida"] is False


def test_update_excel_aisla_ticker_sin_datos(excel_path):
    data = {
        "GGAL.BA": _vela("2026-05-18", 100.0, 110.0, 99.0, 108.0),
        "YPFD.BA": None,  # fallo aislado, no debe romper el Excel
    }

    main.update_excel(data, filename=excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws.cell(row=2, column=6).value == 108.0  # GGAL ok
    # YPFD no se escribió porque el bloque solo tiene la fila de GGAL
